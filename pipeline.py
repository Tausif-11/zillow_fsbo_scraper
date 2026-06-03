import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def parse_and_transform(json_data):
    """Sifts through returned async response layout maps and handles fallback strings cleanly."""
    list_results = json_data.get("cat2", {}).get("searchResults", {}).get("listResults", [])
    print(f"[+] Isolated {len(list_results)} processing data nodes out of stream array.")
    
    records = []
    for item in list_results:
        if not item or not isinstance(item, dict):
            continue
            
        hdp_info = item.get("hdpData", {}).get("homeInfo", {})
        zpid = item.get("zpid") or hdp_info.get("zpid")
        
        detail_url = item.get("detailUrl")
        if detail_url and not detail_url.startswith("http"):
            detail_url = f"https://www.zillow.com{detail_url}"
            
        # Clean Extraction Logic for Price Reductions vs Features
        raw_flex_text = hdp_info.get("priceReduction") or item.get("flexFieldText") or ""
        raw_flex_text = str(raw_flex_text).strip()
        
        price_reduction = "None"
        marketing_features = "None"
        
        if raw_flex_text:
            # Check if the string contains a dollar sign or numeric indicators of a price cut
            if "$" in raw_flex_text or any(char.isdigit() for char in raw_flex_text) and ("day" not in raw_flex_text.lower()):
                price_reduction = raw_flex_text
            else:
                marketing_features = raw_flex_text

        records.append({
            "ZPID": str(zpid) if zpid else "N/A",
            "Price": item.get("price", "N/A"),
            "Unformatted Price": hdp_info.get("price") or item.get("unformattedPrice", 0),
            "Beds": item.get("beds") or hdp_info.get("bedrooms", 0),
            "Baths": item.get("baths") or hdp_info.get("bathrooms", 0),
            "Square Footage": item.get("area") or hdp_info.get("livingArea", "N/A"),
            "Home Type": hdp_info.get("homeType", "N/A"),
            "Address": item.get("address", "N/A"),
            "City": item.get("addressCity") or hdp_info.get("city", "N/A"),
            "State": item.get("addressState") or hdp_info.get("state", "N/A"),
            "Zip Code": item.get("addressZipcode") or hdp_info.get("zipcode", "N/A"),
            "Days on Zillow": hdp_info.get("daysOnZillow", "N/A"),
            "Price Reduction": price_reduction,       # Strictly numeric price cuts
            "Marketing Features": marketing_features, # Catches text like "modern finishes", "waterfront"
            "Detail URL": detail_url or "N/A"
        })
    return pd.DataFrame(records)

def save_styled_excel(df, output_path):
    """Transforms raw pandas mappings into clear corporate tracking layouts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Live Automated FSBO"
    ws.views.sheetView[0].showGridLines = True  
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="111111")
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Commit Column Headers
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26
    
    # Commit Rows & Format Custom Data Categories
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        ws.row_dimensions[r_idx].height = 20
        
        for cell in ws[r_idx]:
            cell.font = data_font
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = zebra_fill
                
            # Alignment rules based on data type mapping
            # (Columns A, C, D, E, F, L are numeric elements)
            if cell.column_letter in ['A', 'C', 'D', 'E', 'F', 'L']:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if str(cell.value).startswith("http"):
                cell.font = Font(name="Segoe UI", size=10, color="1A0DAB", underline="single")
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
        
    wb.save(output_path)
    print(f"[√] Professional Styled Tracker Sheet Saved -> {output_path}")